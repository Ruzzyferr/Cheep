"""MACARİSTAN — GVH Árfigyelő günlük fiyat dökümü okuyucusu.

KAYNAK VE HUKUK: Árfigyelő, Gazdasági Versenyhivatal'ın (GVH — Macar Rekabet
Kurumu) yürüttüğü ZORUNLU fiyat bildirim sistemi. Belirli büyüklüğün üstündeki
perakende zincirleri temel gıda ürünlerinin fiyatlarını her gün bildirmek
zorunda; GVH bunu hem bir JSON API'si hem de her gece yenilenen tek bir XLSX
dökümü olarak kamuya açık yayınlıyor. Kimlik doğrulama, API anahtarı ya da
hız sınırı yok; `robots.txt` yok (site bir SPA, `/robots.txt` HTML döndürüyor);
ziyaretçi şartnamesinde (arfigyelo_rendszer_AFF_latogato.pdf) otomatik erişimi
yasaklayan bir madde YOK — yalnızca marka adlarının ve ürün GÖRSELLERİNİN
ticari kullanımı sınırlanıyor, bu yüzden bu scraper görsel URL'si İLETMİYOR.

DİKKAT — ALAN ADI: doğru adres `arfigyelo.gvh.hu`. Yaygın olarak yazılan
`arfigyelo.gov.hu` DİYE BİR HOST YOK (NXDOMAIN).

FİYAT TABANI: satırlar ZİNCİR bazında, mağaza bazında DEĞİL. Her satır o
zincirin ülke genelindeki minimum ve maksimum fiyatını verir. Ölçüldü
(2026-08-28 dökümü): satırların %94,4'ünde minimum = maksimum, yani fiyat
fiilen ulusal. Farklı olan azınlıkta MİNİMUM yayınlanıyor çünkü:
  • minimum GERÇEKTEN ÖDENEN bir fiyattır — ortalama/orta nokta ise hiçbir
    mağazada geçerli olmayan uydurma bir sayı üretirdi (Hırvatistan'da
    referans mağaza seçmemizin sebebiyle aynı ilke);
  • tüm zincirlere aynı kural uygulandığı için karşılaştırma tarafsız kalır.
Bunlar TESLİMAT değil RAF fiyatlarıdır — Almanya'yı NO-GO yapan
"teslimat platformu fiyatı raf fiyatının %12,5 üstünde" sorunu burada yok.

ÜRÜN KİMLİĞİ: `Termék azonosító` çoğunlukla gerçek EAN. Ama açık/tartılan
ürünlerde (havuç, karnabahar) dolgulu iç kod kullanılıyor ("0000000023023").
Barkod olarak İLETİLİR ama pipeline'ın `is_globally_unique_ean` doğrulaması
geçersizleri düşürür — ölçüm: Auchan %96,4 / Tesco %97,1 geçerli EAN, buna
karşılık indirimcilerde (Aldi %33, Lidl %35, Penny %45) kendi markalı ürünler
iç kod taşıdığı için oran düşük. Geçersiz barkodlu ürün yine içe aktarılır,
yalnızca marketler-arası birleştirmeye girmez.
"""
from __future__ import annotations

import logging
import warnings
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import requests

from countries._common.daily_artifact import fetch_daily
from scrapers.units import parse_quantity_and_unit

logger = logging.getLogger(__name__)

XLSX_URL = "https://cdnarfigyeloprodweu.azureedge.net/excel/arfigyelo_napi_termekadatok.xlsx"
CATEGORIES_URL = "https://arfigyelo.gvh.hu/api/categories"
SHOPS_URL = "https://arfigyelo.gvh.hu/api/shops"
CHAIN_STORES_URL = "https://arfigyelo.gvh.hu/api/chain-stores"

ARTIFACT_NAME = "arfigyelo.xlsx"
CATEGORIES_NAME = "categories.json"

USER_AGENT = "cheep-scraper/1.0 (+https://cheep.live)"

#: Sağlıklı döküm ~1,8 MB. Bunun çok altı, kaynağın 200 ile hata gövdesi
#: döndürdüğü anlamına gelir — önbelleğe yazılmamalı.
MIN_XLSX_BYTES = 300 * 1024

#: XLSX sütun sırası (2026-08-28 itibarıyla doğrulandı). Sıraya değil ADA göre
#: okuyoruz: kaynak araya sütun eklerse indeks tabanlı okuma sessizce YANLIŞ
#: alanı fiyat sanardı.
COL_PRODUCT_ID = "Termék azonosító"
COL_NAME = "Termék név"
COL_CATEGORY_ID = "Kategória azonosító"
COL_CATEGORY_NAME = "Kategória név"
COL_CHAIN = "Üzletlánc név"
COL_UNIT = "Egység"
COL_PACK = "Kiszerelés"
COL_MIN_PRICE = "Minimum ár"
COL_MAX_PRICE = "Maximum ár"

REQUIRED_COLUMNS = (
    COL_PRODUCT_ID, COL_NAME, COL_CATEGORY_ID, COL_CHAIN, COL_MIN_PRICE,
)


@dataclass
class Product:
    name: str
    brand: Optional[str]
    price: float
    quantity: Optional[float]
    unit: Optional[str]
    barcode: Optional[str]
    raw_category: Optional[str]
    sku: str

    def to_dict(self) -> Dict:
        return asdict(self)


# --------------------------------------------------------------- saf yardımcılar

def parse_hu_number(value) -> Optional[float]:
    """Macar ondalık biçimini sayıya çevirir (SAF).

    Döküm fiyatları "499,0000" biçiminde yazıyor — VİRGÜL ondalık ayracı.
    `float("499,0000")` istisna fırlatır; naif bir `float()` çağrısı bu ülkenin
    TÜM fiyatlarını düşürürdü.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "")
    if not text:
        return None
    # Binlik ayracı yok (kaynakta görülmedi); yalnızca ondalık virgülü çevrilir.
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def flatten_category_paths(payload: Dict) -> Dict[str, str]:
    """Kategori ağacını `{yaprak_id: "ust/orta/yaprak"}` sözlüğüne indirger (SAF).

    XLSX yalnızca yaprak ID'si ve yaprak ADI taşıyor; hiyerarşik yol yalnızca
    API'de var. Yolu kullanmak ÖNEMLİ çünkü `category_map.json` `prefix:`
    anahtarlarıyla tüm bir alt ağacı tek satırda eşleyebiliyor — 161 yaprağı
    tek tek yazmak yerine ~42 orta seviye öneki yetiyor.
    """
    out: Dict[str, str] = {}

    def walk(nodes: Iterable[Dict]) -> None:
        for node in nodes or ():
            children = node.get("categoryNodes") or []
            if children:
                walk(children)
            else:
                node_id = node.get("id")
                path = node.get("path")
                if node_id is not None and path:
                    out[str(node_id)] = str(path)

    walk(payload.get("categories") or [])
    return out


def build_products(
    rows: Iterable[Dict[str, object]],
    chain: str,
    category_paths: Dict[str, str],
) -> List[Product]:
    """Döküm satırlarını tek bir zincirin ürünlerine indirger (SAF)."""
    out: List[Product] = []
    for row in rows:
        if str(row.get(COL_CHAIN) or "").strip() != chain:
            continue
        name = str(row.get(COL_NAME) or "").strip()
        if not name:
            continue
        price = parse_hu_number(row.get(COL_MIN_PRICE))
        if price is None or price <= 0:
            continue

        product_id = str(row.get(COL_PRODUCT_ID) or "").strip()
        cat_id = str(row.get(COL_CATEGORY_ID) or "").strip()
        raw_category = category_paths.get(cat_id) or (
            str(row.get(COL_CATEGORY_NAME) or "").strip() or None
        )

        # Gramaj: dökümdeki `Kiszerelés` + `Egység` ("1" + "db", "1,5" + "l").
        # Anlamlı değilse ada düşülür — çoklu paketleri ad üzerinden doğru
        # okuyan ortak ayrıştırıcı kullanılır.
        pack = str(row.get(COL_PACK) or "").strip()
        unit_raw = str(row.get(COL_UNIT) or "").strip()
        qty, unit = parse_quantity_and_unit(f"{pack} {unit_raw}".strip())
        if qty is None or unit in (None, "adet"):
            qty, unit = parse_quantity_and_unit(name)

        out.append(Product(
            name=name,
            # Dökümde ayrı bir marka sütunu YOK. Ad içinden markayı tahmin
            # etmek yanlış birleştirme üretir; None bırakmak dürüst olan.
            brand=None,
            price=price,
            quantity=qty,
            unit=unit,
            barcode=product_id or None,
            raw_category=raw_category,
            sku=f"{chain}:{product_id}" if product_id else f"{chain}:{name[:48]}",
        ))
    return out


def read_xlsx_rows(path: Path) -> List[Dict[str, object]]:
    """XLSX'i sözlük satırlarına çevirir; sütunlar ADA göre okunur."""
    import openpyxl  # yalnızca Macaristan'a özgü bağımlılık, tembel içe aktarım

    with warnings.catch_warnings():
        # openpyxl bu dosya için "no default style" uyarısı basıyor; zararsız
        # ama her koşuda log'u kirletiyor.
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        missing = [c for c in REQUIRED_COLUMNS if c not in header]
        if missing:
            # Kaynak sütun adını değiştirdi. SESSİZ boş liste dönmek, zincirin
            # özetten kaybolmasına ve prune'un tetiklenmesine yol açar.
            raise KeyError(f"Árfigyelő dökümünde beklenen sütunlar yok: {missing} (bulunan: {header})")
        return [dict(zip(header, row)) for row in rows]
    finally:
        wb.close()


# ------------------------------------------------------------------ scraper tabanı

class ArfigyeloChainScraper:
    """Bir Macar zincirinin günlük dökümden okunan fiyatları.

    Alt sınıflar yalnızca `chain` (dökümdeki `Üzletlánc név` değeri) tanımlar.
    Döküm ve kategori ağacı gün başına BİR KEZ indirilir; beş zincir aynı
    dosyayı paylaşır (bkz. countries/_common/daily_artifact.py).
    """

    chain: str = ""

    def __init__(self, cache_dir: Optional[Path] = None, session: Optional[requests.Session] = None):
        if not self.chain:
            raise ValueError(f"{type(self).__name__}: `chain` tanımlanmalı")
        self.cache_dir = Path(cache_dir or Path(__file__).resolve().parents[1] / "cache")
        self.session = session or requests.Session()
        self.session.headers.setdefault("User-Agent", USER_AGENT)

    def _category_paths(self) -> Dict[str, str]:
        path = fetch_daily(
            CATEGORIES_URL, self.cache_dir, CATEGORIES_NAME,
            session=self.session, min_bytes=1024, timeout=60,
        )
        import json
        return flatten_category_paths(json.loads(path.read_text(encoding="utf-8")))

    def fetch_products(self) -> List[Dict]:
        xlsx = fetch_daily(
            XLSX_URL, self.cache_dir, ARTIFACT_NAME,
            session=self.session, min_bytes=MIN_XLSX_BYTES, timeout=300,
        )
        category_paths = self._category_paths()
        rows = read_xlsx_rows(xlsx)
        built = build_products(rows, self.chain, category_paths)
        logger.info("%s: %d ürün (döküm %d satır, %d yaprak kategori)",
                    self.chain, len(built), len(rows), len(category_paths))
        if not built:
            # Zincir dökümden düştü (bildirim kesildi ya da ad değişti).
            # Açık hata, sessiz sıfırdan güvenli.
            raise ValueError(
                f"{self.chain}: dökümde bu zincire ait satır YOK — "
                "`Üzletlánc név` değeri değişmiş olabilir"
            )
        return [p.to_dict() for p in built]


class TescoHuScraper(ArfigyeloChainScraper):
    chain = "Tesco"


class AuchanHuScraper(ArfigyeloChainScraper):
    chain = "Auchan"


class AldiHuScraper(ArfigyeloChainScraper):
    chain = "Aldi"


class LidlHuScraper(ArfigyeloChainScraper):
    chain = "Lidl"


class PennyHuScraper(ArfigyeloChainScraper):
    chain = "Penny"


class DmHuScraper(ArfigyeloChainScraper):
    chain = "dm"


class RossmannHuScraper(ArfigyeloChainScraper):
    chain = "Rossmann"


class MullerHuScraper(ArfigyeloChainScraper):
    chain = "Müller"
